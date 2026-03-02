"""
ARGUS Enterprise Excel Export Parser

Parses ARGUS cash flow exports (.xlsx) into the CRE Lytic deal data model.
Handles the standard ARGUS export sheets:
  - Executive Summary Report
  - Cash Flow
  - NOI
  - Lease Summary Report
  - All Terms Report
  - Market Leasing Summary
  - Building Area Occupancy Report
"""

import re
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


def is_argus_file(excel_path: str) -> bool:
    """Detect if an Excel file is an ARGUS export by checking sheet names."""
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = [s.lower().strip() for s in wb.sheetnames]
        wb.close()
        argus_indicators = ["cash flow", "executive summary report", "lease summary report", "noi"]
        matches = sum(1 for ind in argus_indicators if any(ind in sn for sn in sheet_names))
        return matches >= 2
    except Exception:
        return False


def _find_sheet(wb, keywords: List[str]):
    """Find a sheet by partial name match."""
    for name in wb.sheetnames:
        lower = name.lower().strip()
        for kw in keywords:
            if kw in lower:
                return wb[name]
    return None


def _cell_val(ws, row, col):
    """Get cell value, returning None for empty."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == '' or val == 'NaN':
            return None
    return val


def _to_float(val) -> float:
    """Convert a value to float, handling strings with commas/parens."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(',', '').replace('$', '').replace('%', '').strip()
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def parse_executive_summary(wb) -> Dict[str, Any]:
    """Parse the Executive Summary Report sheet."""
    ws = _find_sheet(wb, ["executive summary"])
    if not ws:
        return {}

    result = {}
    for row in range(1, ws.max_row + 1):
        label = _cell_val(ws, row, 1)
        value = _cell_val(ws, row, 2)
        if not label:
            continue
        label_lower = str(label).lower().strip()

        if label_lower == "property name":
            result["name"] = str(value)
        elif label_lower == "property type":
            result["property_type"] = str(value)
        elif label_lower == "building area":
            result["total_sf"] = _to_float(value)
        elif "acquisition" in label_lower or "valuation date" in label_lower:
            result["acquisition_date"] = str(value)
        elif label_lower == "general inflation":
            result["inflation"] = _to_float(value)
        elif "capital expenditures" in label_lower:
            result["capex_psf"] = _to_float(value)
        elif "general vacancy" in label_lower:
            result["general_vacancy"] = _to_float(value)
        elif "net operating income" in label_lower:
            # Parse "918,072 (52.50%)" format
            if value:
                match = re.search(r'([\d,]+)', str(value))
                if match:
                    result["noi_year1"] = _to_float(match.group(1))
        elif "effective gross revenue" in label_lower and "%" in str(value or ""):
            match = re.search(r'([\d,]+)', str(value))
            if match:
                result["egr_year1"] = _to_float(match.group(1))
        elif "operating expenses" in label_lower and "%" in str(value or ""):
            match = re.search(r'([\d,]+)', str(value))
            if match:
                result["opex_year1"] = _to_float(match.group(1))
        elif "cash flow before debt" in label_lower:
            result["cfbds_year1"] = _to_float(value)
        elif "total leasing" in label_lower:
            result["leasing_costs_year1"] = _to_float(value)
        elif "occupancy" in label_lower and "nra" in label_lower:
            # Parse "42,443 / 94.62%"
            if value:
                match = re.search(r'([\d,.]+)%', str(value))
                if match:
                    result["occupancy_pct"] = _to_float(match.group(1)) / 100.0
                match_sf = re.search(r'([\d,]+)', str(value))
                if match_sf:
                    result["occupied_sf"] = _to_float(match_sf.group(1))
        elif label_lower == "walt":
            result["walt"] = str(value)
        elif "available sf" in label_lower:
            result["available_sf"] = _to_float(value)

    return result


def parse_cash_flow(wb) -> Dict[str, Any]:
    """Parse the Cash Flow sheet — multi-year projections."""
    ws = _find_sheet(wb, ["cash flow"])
    if not ws:
        return {}

    # Find the header row with "Year 1", "Year 2", etc.
    year_row = None
    label_col = 1
    for row in range(1, min(ws.max_row + 1, 15)):
        for col in range(1, ws.max_column + 1):
            val = _cell_val(ws, row, col)
            if val and str(val).strip() == "Year 1":
                year_row = row
                break
        if year_row:
            break

    if not year_row:
        return {}

    # Count years
    year_cols = {}
    for col in range(1, ws.max_column + 1):
        val = _cell_val(ws, year_row, col)
        if val and str(val).startswith("Year"):
            try:
                yr_num = int(str(val).replace("Year", "").strip())
                year_cols[yr_num] = col
            except ValueError:
                pass

    # Also check for per-SF column (column 2 often has $/SF values)
    psf_col = None
    for col in range(1, ws.max_column + 1):
        val = _cell_val(ws, year_row - 1, col)
        if val and "psf" in str(val).lower():
            psf_col = col

    # Parse rows by label
    cash_flow_data = {"years": {}, "line_items": {}}
    num_years = len(year_cols)

    for row in range(year_row + 1, ws.max_row + 1):
        label = _cell_val(ws, row, label_col)
        if not label or isinstance(label, (int, float)):
            continue
        label_str = str(label).strip()
        if not label_str:
            continue

        row_values = {}
        for yr_num, col in year_cols.items():
            val = _cell_val(ws, row, col)
            if val is not None:
                row_values[yr_num] = _to_float(val)

        if row_values:
            cash_flow_data["line_items"][label_str] = row_values

    # Extract key metrics
    key_map = {
        "Potential Base Rent": "base_rent",
        "Total Rental Revenue": "total_rental_revenue",
        "Total Expense Recoveries": "expense_recoveries",
        "Effective Gross Revenue": "egr",
        "Total Operating Expenses": "total_opex",
        "Net Operating Income": "noi",
        "Total Leasing & Capital Costs": "total_leasing_costs",
        "Cash Flow Before Debt Service": "cfbds",
        "Tenant Improvements": "tenant_improvements",
        "Leasing Commissions": "leasing_commissions",
        "Capital Reserves": "capital_reserves",
    }

    for yr_num in sorted(year_cols.keys()):
        year_data = {}
        for argus_label, field_name in key_map.items():
            if argus_label in cash_flow_data["line_items"]:
                year_data[field_name] = cash_flow_data["line_items"][argus_label].get(yr_num, 0)
        cash_flow_data["years"][yr_num] = year_data

    # Extract individual OpEx line items for Year 1
    opex_items = {}
    in_opex = False
    for row in range(year_row + 1, ws.max_row + 1):
        label = _cell_val(ws, row, label_col)
        if not label:
            continue
        label_str = str(label).strip()
        if label_str == "Operating Expenses":
            in_opex = True
            continue
        if label_str == "Total Operating Expenses":
            in_opex = False
            continue
        if in_opex and label_str:
            yr1_col = year_cols.get(1)
            if yr1_col:
                val = _cell_val(ws, row, yr1_col)
                if val is not None:
                    opex_items[label_str] = _to_float(val)

    cash_flow_data["opex_detail"] = opex_items
    cash_flow_data["num_years"] = num_years

    return cash_flow_data


def parse_lease_summary(wb) -> List[Dict[str, Any]]:
    """Parse the Lease Summary Report into a rent roll."""
    ws = _find_sheet(wb, ["lease summary"])
    if not ws:
        return []

    rent_roll = []
    row = 1
    max_row = ws.max_row

    # Skip header rows — find first tenant (numbered "1. TenantName")
    while row <= max_row:
        val = _cell_val(ws, row, 1)
        if val and re.match(r'^\d+\.', str(val)):
            break
        row += 1

    # Parse tenants — each tenant spans 4 rows:
    # Row 1: "N. Tenant Name" | SF | Rate/SF/Year | Escalation date | Escalation rate | Recovery | Space Type
    # Row 2: "Suite: XXX" | Building Share % | Amount/Year | Escalation date | Escalation rate | ...
    # Row 3: "Start - End" | ... | Rate/Month | ...
    # Row 4: "Term Length" | ... | Amount/Month | ...
    # Row 5: blank separator

    while row <= max_row:
        val = _cell_val(ws, row, 1)
        if val is None:
            row += 1
            continue

        val_str = str(val).strip()

        # Check if this is a tenant row (starts with "N. Name")
        tenant_match = re.match(r'^(\d+)\.\s*(.+)', val_str)
        if not tenant_match:
            row += 1
            continue

        tenant_num = tenant_match.group(1)
        tenant_name = tenant_match.group(2).strip()

        # Parse the 4 rows of tenant data
        sf = _to_float(_cell_val(ws, row, 2))
        rent_psf = _to_float(_cell_val(ws, row, 3))
        space_type = _cell_val(ws, row, 7) or ""
        recovery_type = _cell_val(ws, row, 6) or ""

        # Row 2: Suite and annual rent
        suite = ""
        annual_rent = 0
        building_share = 0
        if row + 1 <= max_row:
            suite_val = _cell_val(ws, row + 1, 1)
            if suite_val:
                suite_match = re.match(r'Suite:\s*(.+)', str(suite_val))
                if suite_match:
                    suite = suite_match.group(1).strip()
            building_share = _to_float(_cell_val(ws, row + 1, 2))
            annual_rent = _to_float(_cell_val(ws, row + 1, 3))

        # Row 3: Lease dates
        lease_start = None
        lease_end = None
        if row + 2 <= max_row:
            date_val = _cell_val(ws, row + 2, 1)
            if date_val:
                date_match = re.match(r'(\d+/\d+/\d+)\s*-\s*(\d+/\d+/\d+)', str(date_val))
                if date_match:
                    lease_start = date_match.group(1)
                    lease_end = date_match.group(2)

        # Row 4: Term length
        term_length = ""
        if row + 3 <= max_row:
            term_val = _cell_val(ws, row + 3, 1)
            if term_val:
                term_length = str(term_val).strip()

        # Collect rent escalations from columns 4-5
        escalations = []
        for esc_row in range(row, min(row + 4, max_row + 1)):
            esc_date = _cell_val(ws, esc_row, 4)
            esc_rate = _cell_val(ws, esc_row, 5)
            if esc_date and esc_rate:
                escalations.append({
                    "date": str(esc_date),
                    "rate_psf": _to_float(esc_rate),
                })

        rent_roll.append({
            "unit": suite,
            "tenant": tenant_name,
            "sf": sf,
            "rent_psf": rent_psf,
            "annual_rent": annual_rent,
            "lease_start": lease_start,
            "lease_end": lease_end,
            "term": term_length,
            "space_type": str(space_type),
            "recovery_type": str(recovery_type),
            "building_share_pct": building_share,
            "escalations": escalations,
        })

        # Skip to next tenant (advance past the 4-5 rows)
        row += 5

    return rent_roll


def parse_all_terms(wb) -> Dict[str, Any]:
    """Parse All Terms Report — lease expiration schedule."""
    ws = _find_sheet(wb, ["all terms"])
    if not ws:
        return {}

    # Find year header row
    year_row = None
    for row in range(1, min(ws.max_row + 1, 15)):
        for col in range(1, ws.max_column + 1):
            val = _cell_val(ws, row, col)
            if val and str(val).strip() == "Year 1":
                year_row = row
                break
        if year_row:
            break

    if not year_row:
        return {}

    # Get year columns
    year_cols = {}
    for col in range(1, ws.max_column + 1):
        val = _cell_val(ws, year_row, col)
        if val and str(val).startswith("Year"):
            try:
                yr_num = int(str(val).replace("Year", "").strip())
                year_cols[yr_num] = col
            except ValueError:
                pass

    # Parse expiration schedule
    expirations = {}
    for row in range(year_row + 1, ws.max_row + 1):
        label = _cell_val(ws, row, 1)
        if not label:
            continue
        label_str = str(label).strip()
        if "total sf expiring" in label_str.lower():
            for yr_num, col in year_cols.items():
                val = _cell_val(ws, row, col)
                if val is not None:
                    expirations[yr_num] = _to_float(val)
            break

    return {"lease_expirations_sf": expirations}


def parse_market_leasing(wb) -> Dict[str, Any]:
    """Parse Market Leasing Summary."""
    ws = _find_sheet(wb, ["market leasing"])
    if not ws:
        return {}

    market_assumptions = {}
    current_space_type = None

    for row in range(1, ws.max_row + 1):
        # Space types appear in specific columns
        for col in [3, 5]:
            val = _cell_val(ws, row, col)
            if val and str(val).strip() in ["Laboratory", "Office", "Retail", "Industrial", "Warehouse"]:
                current_space_type = str(val).strip()
                market_assumptions[current_space_type] = {}

        label = _cell_val(ws, row, 1)
        if not label:
            continue
        label_str = str(label).strip().lower()

        for col, space_type in [(3, None), (5, None)]:
            val = _cell_val(ws, row, col)
            if val is None:
                continue
            # Figure out which space type this column belongs to
            for st_name, st_data in market_assumptions.items():
                if "market base rent" in label_str:
                    st_data["market_rent_psf"] = _to_float(val)
                elif "rent increase" in label_str:
                    pct = str(val).replace('%', '').strip()
                    st_data["rent_increase_pct"] = _to_float(pct) / 100.0 if _to_float(pct) > 1 else _to_float(pct)
                elif "term length" in label_str:
                    st_data["term_length"] = str(val)
                elif "renewal probability" in label_str:
                    st_data["renewal_probability"] = _to_float(val)
                elif "months vacant" in label_str:
                    st_data["months_vacant"] = _to_float(val)
                elif "tenant improvements" in label_str:
                    st_data["ti_psf"] = _to_float(val)
                elif "free rent" in label_str:
                    st_data["free_rent_months"] = _to_float(val)
                elif "leasing commissions" in label_str:
                    st_data["leasing_commission_pct"] = _to_float(val)

    return market_assumptions


def parse_argus_file(excel_path: str) -> Dict[str, Any]:
    """
    Main entry point: parse an ARGUS Excel export into the CRE Lytic deal data model.

    Returns the same structure as parse_offering_memorandum() so the dashboard
    can consume it seamlessly.
    """
    wb = load_workbook(excel_path, data_only=True)

    exec_summary = parse_executive_summary(wb)
    cash_flow = parse_cash_flow(wb)
    rent_roll = parse_lease_summary(wb)
    all_terms = parse_all_terms(wb)
    market_leasing = parse_market_leasing(wb)

    wb.close()

    # Build the canonical deal structure
    total_sf = exec_summary.get("total_sf", 0)
    noi = 0
    cap_rate = 0

    # Get Year 1 NOI from cash flow
    if cash_flow.get("years", {}).get(1):
        noi = cash_flow["years"][1].get("noi", 0)
    elif exec_summary.get("noi_year1"):
        noi = exec_summary["noi_year1"]

    # Calculate annual rent from rent roll
    total_annual_rent = sum(t.get("annual_rent", 0) for t in rent_roll)

    # NOI growth from Year 1 to Year 2
    noi_growth = 0.02  # default
    if cash_flow.get("years", {}).get(1) and cash_flow.get("years", {}).get(2):
        y1_noi = cash_flow["years"][1].get("noi", 0)
        y2_noi = cash_flow["years"][2].get("noi", 0)
        if y1_noi > 0:
            noi_growth = (y2_noi - y1_noi) / y1_noi

    # Build multi-year cash flow projections for the dashboard
    cf_projections = []
    for yr_num in sorted(cash_flow.get("years", {}).keys()):
        yr = cash_flow["years"][yr_num]
        cf_projections.append({
            "year": yr_num,
            "noi": yr.get("noi", 0),
            "egr": yr.get("egr", 0),
            "total_opex": yr.get("total_opex", 0),
            "cfbds": yr.get("cfbds", 0),
            "tenant_improvements": yr.get("tenant_improvements", 0),
            "leasing_commissions": yr.get("leasing_commissions", 0),
            "capital_reserves": yr.get("capital_reserves", 0),
        })

    parsed_data = {
        "source": "argus",
        "property": {
            "name": exec_summary.get("name", "ARGUS Import"),
            "property_type": exec_summary.get("property_type", "N/A"),
            "total_sf": total_sf,
            "address": "",
            "city": "",
            "state": "",
            "asking_price": 0,  # ARGUS doesn't have asking price — user sets via cap rate
            "occupancy_pct": exec_summary.get("occupancy_pct", 0),
            "occupied_sf": exec_summary.get("occupied_sf", 0),
            "available_sf": exec_summary.get("available_sf", 0),
            "walt": exec_summary.get("walt", ""),
            "acquisition_date": exec_summary.get("acquisition_date", ""),
        },
        "financials": {
            "noi": noi,
            "egr": exec_summary.get("egr_year1", 0),
            "total_opex": exec_summary.get("opex_year1", 0),
            "cfbds": exec_summary.get("cfbds_year1", 0),
            "cap_rate": 0,  # No asking price from ARGUS — cap rate derived from user input
            "total_sf": total_sf,
        },
        "assumptions": {
            "exit_cap_rate": 0.065,
            "noi_growth": round(noi_growth, 4),
            "hold_period": min(cash_flow.get("num_years", 10), 10),
            "ltv": 0.65,
            "interest_rate": 0.055,
            "amortization_years": 25,
            "general_inflation": exec_summary.get("inflation", 0.03),
            "general_vacancy": exec_summary.get("general_vacancy", 0.035),
        },
        "rent_roll": [{
            "unit": t.get("unit", ""),
            "tenant": t.get("tenant", ""),
            "sf": t.get("sf", 0),
            "rent_psf": t.get("rent_psf", 0),
            "annual_rent": t.get("annual_rent", 0),
            "lease_start": t.get("lease_start"),
            "lease_end": t.get("lease_end"),
            "term": t.get("term", ""),
            "space_type": t.get("space_type", ""),
            "recovery_type": t.get("recovery_type", ""),
            "escalations": t.get("escalations", []),
        } for t in rent_roll],
        "argus_detail": {
            "cash_flow_projections": cf_projections,
            "opex_detail": cash_flow.get("opex_detail", {}),
            "lease_expirations": all_terms.get("lease_expirations_sf", {}),
            "market_leasing": market_leasing,
        },
    }

    parsing_report = {
        "source": "argus",
        "errors": [],
        "warnings": [],
        "fields_extracted": [],
    }

    if exec_summary.get("name"):
        parsing_report["fields_extracted"].append("property_name")
    if noi > 0:
        parsing_report["fields_extracted"].append("noi")
    if total_sf > 0:
        parsing_report["fields_extracted"].append("total_sf")
    if rent_roll:
        parsing_report["fields_extracted"].append(f"rent_roll ({len(rent_roll)} tenants)")
    if cf_projections:
        parsing_report["fields_extracted"].append(f"cash_flow ({len(cf_projections)} years)")
    if not noi:
        parsing_report["warnings"].append("No NOI found in ARGUS export")
    if not rent_roll:
        parsing_report["warnings"].append("No rent roll found in Lease Summary Report")

    return {
        "parsed_data": parsed_data,
        "parsing_report": parsing_report,
    }
