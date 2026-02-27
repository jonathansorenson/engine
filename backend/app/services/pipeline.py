import re
import json
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path
from difflib import SequenceMatcher
import pdfplumber
from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════════
# FUZZY COLUMN MATCHING — handles varied header names
# ═══════════════════════════════════════════════════════════════

# Each canonical field maps to a list of possible header variations
COLUMN_ALIASES = {
    "unit": [
        "unit", "suite", "space", "unit #", "unit no", "suite #", "suite no",
        "space #", "space no", "unit number", "suite number", "bldg", "building",
        "unit/suite", "ste", "spc", "bay", "pad", "shop", "store", "floor",
    ],
    "tenant": [
        "tenant", "tenant name", "occupant", "lessee", "company", "name",
        "business", "client", "renter", "entity", "leaseholder", "party",
        "tenant/occupant", "current tenant", "tenant name/dba",
    ],
    "sf": [
        "sf", "sq ft", "sqft", "square feet", "square footage", "area",
        "rsf", "rentable sf", "usable sf", "nra", "gla", "size",
        "rentable area", "net rentable", "leased sf", "occupied sf",
        "total sf", "approx sf", "approx. sf", "leasable area",
        "actual sq ft", "lease sq ft", "useable sq ft",
    ],
    "annual_rent": [
        "annual rent", "annual base rent", "base rent", "yearly rent",
        "annual", "total rent", "contract rent", "scheduled rent",
        "annual base", "ann rent", "ann. rent", "rent/yr", "rent/year",
        "annual income", "gross rent", "year rent", "current year rent",
    ],
    "monthly_rent": [
        "monthly rent", "monthly base", "rent/mo", "rent/month",
        "mo rent", "mo. rent", "monthly", "mthly rent", "month rent",
        "current monthly rent", "total monthly income",
    ],
    "rent_psf": [
        "rent psf", "psf", "$/sf", "rent/sf", "rate", "rate/sf",
        "per sf", "per sq ft", "rent per sf", "base rate", "nnn rate",
        "contract rate", "asking rate", "effective rate", "price/sf",
        "rent per sq ft", "current rent per sq ft", "total psf",
    ],
    "lease_start": [
        "lease start", "start date", "commencement", "commence date",
        "move in", "lease commencement", "begin", "begin date",
        "effective date", "lease from", "occ date",
        "original lease start date", "lease start date",
    ],
    "lease_end": [
        "lease end", "end date", "expiration", "expiry", "expire",
        "expiry date", "expiration date", "lease expiration", "term end",
        "lease to", "lease thru", "maturity", "termination",
        "lease expir date", "lease expiration date",
    ],
    "lease_type": [
        "lease type", "lease class", "structure",
        "nnn", "gross", "modified gross", "full service",
    ],
    "status": [
        "status", "occupied", "occupancy", "vacancy", "vacant",
        "occ status", "lease status", "current status",
    ],
}

# Short/ambiguous headers that should NOT match on their own via substring
# These cause false positives (e.g. "Rent" matching annual_rent when it's monthly)
AMBIGUOUS_HEADERS = {"rent", "income", "type", "class", "start", "from", "to", "lease"}


def fuzzy_match_column(header: str) -> Optional[str]:
    """Match a column header to a canonical field name using fuzzy matching."""
    if not header:
        return None

    h = str(header).lower().strip()
    h = re.sub(r'[^\w\s/\$#]', '', h).strip()

    if not h or len(h) < 2:
        return None

    best_field = None
    best_score = 0.0

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            # Exact match — highest priority
            if h == alias:
                return field

        # Fuzzy matching (only if no exact match found above)
        for alias in aliases:
            # Contains match (alias is substring of header or vice versa)
            if len(h) > 3 and (alias in h or h in alias):
                score = len(alias) / max(len(h), len(alias))
                # Penalize very short headers matching long aliases via substring
                if len(h) <= 4 and h in AMBIGUOUS_HEADERS:
                    score *= 0.3  # Heavy penalty for ambiguous short headers
                if score > best_score and score > 0.4:
                    best_score = score
                    best_field = field

            # Fuzzy ratio
            ratio = SequenceMatcher(None, h, alias).ratio()
            if ratio > best_score and ratio > 0.7:
                best_score = ratio
                best_field = field

    return best_field


# ═══════════════════════════════════════════════════════════════
# EXCEL PARSER — format-agnostic with multi-row header support
# ═══════════════════════════════════════════════════════════════

def _is_header_text(cell) -> bool:
    """Check if a cell value looks like a header (not a date, number, or junk)."""
    if cell is None:
        return False
    from datetime import datetime
    if isinstance(cell, datetime):
        return False
    cell_str = str(cell).strip()
    if not cell_str or len(cell_str) > 40:
        return False
    # Skip pure numbers, dates, percentages
    if cell_str.replace('.', '').replace(',', '').replace('-', '').replace('/', '').isdigit():
        return False
    # Skip date-like strings
    if re.match(r'\d{4}-\d{2}-\d{2}', cell_str):
        return False
    # Skip things that look like "add on factor" or notes
    lower = cell_str.lower()
    if any(skip in lower for skip in ['factor', 'note', 'formula', 'total', 'subtotal']):
        return False
    return True


def _build_merged_headers(rows: List[List], candidate_idx: int) -> List[str]:
    """
    Build merged column headers by concatenating up to 3 rows above + the candidate row.
    This handles multi-row headers like:
        Row 7: [Current, Current, ...]
        Row 8: [Monthly, Monthly, Year, ...]
        Row 9: [Tenant, Suite #, Sq. Ft., Rent, ...]
    → merged: [Tenant, Suite #, Sq. Ft., Current Monthly Rent, ...]
    """
    ncols = len(rows[candidate_idx]) if candidate_idx < len(rows) else 0
    merged = [""] * ncols

    # Look at up to 3 rows above the candidate for sub-headers
    start = max(0, candidate_idx - 3)
    for row_idx in range(start, candidate_idx + 1):
        if row_idx >= len(rows):
            continue
        row = rows[row_idx]
        for ci in range(min(len(row), ncols)):
            cell = row[ci]
            if _is_header_text(cell):
                cell_str = str(cell).strip()
                if merged[ci]:
                    merged[ci] = merged[ci] + " " + cell_str
                else:
                    merged[ci] = cell_str

    return merged


def extract_excel_rent_roll(excel_path: str) -> Optional[List[Dict[str, Any]]]:
    """
    Extract rent roll from Excel file.

    Strategy:
    1. Scan every sheet
    2. For each sheet, scan rows to find the best header row (most recognized columns)
    3. Merge multi-row headers by looking at rows above the candidate
    4. Parse all data rows below the header
    5. Return the best result (most rows with most fields)
    """
    all_results = []

    try:
        workbook = load_workbook(excel_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Collect all rows
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))

            if not rows or len(rows) < 2:
                continue

            # === STRATEGY 1: Find header row with multi-row merge ===
            header_row_idx = None
            best_col_map = {}

            for idx in range(min(25, len(rows))):
                row = rows[idx]
                if not row or not any(row):
                    continue

                # Count string cells (potential header row has mostly strings)
                str_count = sum(1 for c in row if c and isinstance(c, str))
                if str_count < 2:
                    continue

                # Try matching with just this row's values
                col_map_single = {}
                for ci, cell in enumerate(row):
                    if cell is None:
                        continue
                    field = fuzzy_match_column(str(cell))
                    if field and field not in col_map_single.values():
                        col_map_single[ci] = field

                # Also try with merged multi-row headers
                merged_headers = _build_merged_headers(rows, idx)
                col_map_merged = {}
                for ci, header_text in enumerate(merged_headers):
                    if not header_text:
                        continue
                    field = fuzzy_match_column(header_text)
                    if field and field not in col_map_merged.values():
                        col_map_merged[ci] = field

                # Use whichever approach found more columns
                col_map = col_map_merged if len(col_map_merged) >= len(col_map_single) else col_map_single

                # Need at least 2 recognized columns to consider it a header
                if len(col_map) >= 2 and len(col_map) > len(best_col_map):
                    header_row_idx = idx
                    best_col_map = col_map

            if header_row_idx is None or not best_col_map:
                continue

            # === Fix empty columns: if a mapped column has no data, look for alternatives ===
            # Collect all possible field→column mappings (including duplicates)
            all_field_cols = {}  # field → [col_idx, ...]
            merged_headers = _build_merged_headers(rows, header_row_idx)
            for ci, header_text in enumerate(merged_headers):
                if not header_text:
                    continue
                field = fuzzy_match_column(header_text)
                if field:
                    all_field_cols.setdefault(field, []).append(ci)
            # Also check single-row headers
            for ci, cell in enumerate(rows[header_row_idx]):
                if cell is None:
                    continue
                field = fuzzy_match_column(str(cell))
                if field:
                    all_field_cols.setdefault(field, []).append(ci)

            # For fields with multiple candidate columns, pick the one with most data
            data_rows_sample = rows[header_row_idx + 1: min(header_row_idx + 15, len(rows))]
            for field, col_indices in all_field_cols.items():
                unique_cols = list(dict.fromkeys(col_indices))  # dedupe preserving order
                if len(unique_cols) <= 1:
                    continue
                # Count non-empty cells in each candidate column
                best_ci = unique_cols[0]
                best_count = 0
                for ci in unique_cols:
                    count = sum(1 for r in data_rows_sample
                                if r and ci < len(r) and r[ci] is not None
                                and str(r[ci]).strip() not in ["", "None", "-"])
                    if count > best_count:
                        best_count = count
                        best_ci = ci
                # Update col_map: remove any existing mapping for this field, add best
                best_col_map = {k: v for k, v in best_col_map.items() if v != field}
                best_col_map[best_ci] = field

            # === Parse data rows ===
            rent_roll = []
            for row_idx in range(header_row_idx + 1, len(rows)):
                row = rows[row_idx]
                if not row or not any(row):
                    continue

                # Skip rows that look like totals, subtotals, notes, or factor rows
                first_val = str(row[0] or '').lower().strip() if row else ''
                skip_patterns = ['total', 'subtotal', 'sum', 'grand total', '---',
                                 'add on', 'factor', 'note:', 'notes:', '*']
                if any(skip in first_val for skip in skip_patterns):
                    continue

                entry = {}
                for ci, value in enumerate(row):
                    if ci not in best_col_map or value is None:
                        continue

                    field = best_col_map[ci]
                    val_str = str(value).strip()

                    if not val_str or val_str.lower() in ["", "none", "-", "n/a", "null"]:
                        continue

                    if field in ["sf", "annual_rent", "monthly_rent", "rent_psf"]:
                        try:
                            cleaned = re.sub(r'[,$\s\(\)]', '', val_str)
                            if cleaned and cleaned.replace('.', '').replace('-', '').isdigit():
                                num_val = abs(float(cleaned))
                                # Sanity: SF should be > 50, rents should be > 0
                                if field == "sf" and num_val < 50:
                                    continue
                                entry[field] = num_val
                        except (ValueError, TypeError):
                            pass
                    elif field in ["lease_start", "lease_end"]:
                        try:
                            from datetime import datetime
                            if isinstance(value, datetime):
                                entry[field] = value.strftime("%Y-%m-%d")
                            else:
                                entry[field] = val_str
                        except Exception:
                            entry[field] = val_str
                    else:
                        entry[field] = val_str

                # Derive missing fields
                if "monthly_rent" in entry and "annual_rent" not in entry:
                    entry["annual_rent"] = entry["monthly_rent"] * 12
                if "annual_rent" in entry and "monthly_rent" not in entry:
                    entry["monthly_rent"] = entry["annual_rent"] / 12
                if "annual_rent" in entry and "sf" in entry and "rent_psf" not in entry and entry["sf"] > 0:
                    entry["rent_psf"] = round(entry["annual_rent"] / entry["sf"], 2)
                if "rent_psf" in entry and "sf" in entry and "annual_rent" not in entry:
                    entry["annual_rent"] = entry["rent_psf"] * entry["sf"]

                # STRICT: require at least one identifier AND one numeric value
                has_id = "unit" in entry or "tenant" in entry
                has_num = "sf" in entry or "annual_rent" in entry or "monthly_rent" in entry or "rent_psf" in entry
                if has_id and has_num:
                    rent_roll.append(entry)

            if rent_roll:
                all_results.append({
                    "sheet": sheet_name,
                    "col_map": {k: v for k, v in best_col_map.items()},
                    "data": rent_roll,
                    "score": len(rent_roll) * len(best_col_map),
                })

        workbook.close()

    except Exception as e:
        print(f"Error extracting Excel: {e}")
        import traceback
        traceback.print_exc()

    if not all_results:
        return None

    # Return the best result (highest score)
    best = max(all_results, key=lambda r: r["score"])
    print(f"Excel parser: found {len(best['data'])} rows from sheet '{best['sheet']}' with columns: {list(best['col_map'].values())}")
    return best["data"]


# ═══════════════════════════════════════════════════════════════
# PDF TEXT EXTRACTION
# ═══════════════════════════════════════════════════════════════

def extract_text_from_pdf(pdf_path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """Extract text and tables from PDF using pdfplumber."""
    full_text = ""
    tables = []

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                full_text += page_text + "\n"

                page_tables = page.extract_tables()
                if page_tables:
                    for table in page_tables:
                        if table and len(table) >= 2:
                            tables.append({
                                "data": table,
                                "page": page.page_number,
                            })
    except Exception as e:
        print(f"Error extracting PDF: {e}")
        import traceback
        traceback.print_exc()

    return full_text, tables


# ═══════════════════════════════════════════════════════════════
# PDF PROPERTY & FINANCIAL EXTRACTION — expanded patterns
# ═══════════════════════════════════════════════════════════════

def extract_numeric_value(text: str, pattern: str) -> Optional[float]:
    """Extract numeric value from text using regex pattern."""
    try:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            numbers = re.findall(r"[\d,]+\.?\d*", match.group(0))
            if numbers:
                return float(numbers[0].replace(",", ""))
    except Exception:
        pass
    return None


def find_all_dollar_amounts(text: str) -> List[Dict[str, Any]]:
    """Find all dollar amounts in text with their context."""
    amounts = []
    for match in re.finditer(r'(?:^|\s)([^\n]{0,60}?\$[\d,]+(?:\.\d{2})?(?:\s*(?:M|MM|million|Million|B|billion))?)', text, re.MULTILINE):
        context = match.group(1).strip()
        nums = re.findall(r'\$([\d,]+(?:\.\d{2})?)\s*(?:(M|MM|million|Million|B|billion))?', context)
        if nums:
            val = float(nums[0][0].replace(",", ""))
            multiplier = nums[0][1].lower() if nums[0][1] else ""
            if multiplier in ["m", "mm", "million"]:
                val *= 1_000_000
            elif multiplier in ["b", "billion"]:
                val *= 1_000_000_000
            amounts.append({"value": val, "context": context.lower()})
    return amounts


def extract_property_details(text: str) -> Dict[str, Any]:
    """Extract property details from text with expanded pattern matching."""
    details = {}

    # Use first ~12K chars for property-level details (avoid sponsor bios at end)
    property_text = text[:12000]
    text_lower = text.lower()
    prop_text_lower = property_text.lower()

    # ── Property name ──
    # Strategy 1: Look for "Property <name>" or explicit label
    name_patterns = [
        # "Property Rialto Place Office Tower" format
        r"Property\s+([A-Z][A-Za-z\s&\-\'\.]+(?:Tower|Center|Plaza|Building|Park|Complex|Place|Court))",
        # "Property Name: X" or "Property: X" format
        r"(?:Property\s*Name|Subject\s*Property)[\s:]+([^\n]{3,60})",
    ]
    for pattern in name_patterns:
        name = re.search(pattern, property_text, re.MULTILINE)
        if name:
            candidate = name.group(1).strip()
            # Filter out section headings and garbage
            skip_words = ["overview", "summary", "description", "details", "type", "address"]
            if (len(candidate) > 3 and not candidate.startswith("$")
                    and not candidate[0].isdigit()
                    and not any(sw in candidate.lower() for sw in skip_words)):
                details["name"] = candidate[:60]
                break

    # Strategy 2: First line of document often has property name
    if "name" not in details:
        first_line = text.split("\n")[0].strip()
        if first_line and "|" in first_line:
            # "100 Rialto Place | Melbourne, FL" format
            details["name"] = first_line.split("|")[0].strip()[:60]
        elif first_line and 5 < len(first_line) < 60 and not first_line.startswith("$"):
            details["name"] = first_line[:60]

    # ── Address ──
    address_patterns = [
        # Explicit label patterns (but NOT section headings)
        r"(?:Property\s*Address|Site\s*Address|Located\s*at)[\s:,]+(\d{1,6}\s+[^\n]{5,80})",
        # Street number + name + suffix + optional city/state/zip
        r"(\d{1,6}\s+[A-Z][A-Za-z\s\.]+(?:Street|St|Avenue|Ave|Boulevard|Blvd|Road|Rd|Drive|Dr|Lane|Ln|Way|Circle|Cir|Court|Ct|Place|Pl|Parkway|Pkwy|Highway|Hwy)(?:\s*,?\s*[A-Z][a-z]+(?:\s[A-Z][a-z]+)*\s*,?\s*[A-Z]{2}\s*\d{5})?)",
    ]
    for pattern in address_patterns:
        addr = re.search(pattern, property_text, re.IGNORECASE | re.MULTILINE)
        if addr:
            candidate = addr.group(1).strip()
            # Make sure it's not a section heading
            if candidate.lower() not in ["overview", "summary", "description"]:
                # Trim after zip code if present
                zip_match = re.search(r'(\d{5}(?:-\d{4})?)', candidate)
                if zip_match:
                    candidate = candidate[:zip_match.end()].strip()
                # Remove trailing quotes/parens
                candidate = re.sub(r'[\s("\']+$', '', candidate)
                details["address"] = candidate[:100]
                break

    # ── City, State, Zip ──
    csz_pattern = r"([A-Z][a-z]+(?:\s[A-Z][a-z]+)*),?\s+([A-Z]{2})\s+(\d{5}(?:-\d{4})?)"
    csz = re.search(csz_pattern, property_text)
    if csz:
        details["city"] = csz.group(1)
        details["state"] = csz.group(2)
        details["zip"] = csz.group(3)
        if "address" not in details:
            details["address"] = f"{csz.group(1)}, {csz.group(2)} {csz.group(3)}"

    # ── Property type ──
    # IMPORTANT: Check more specific types FIRST to avoid false positives
    # Only search the first ~12K chars to avoid sponsor bios mentioning other types
    type_keywords_ordered = [
        ("office", ["office building", "office tower", "class a office", "class b office",
                     "office", "office space", "professional office"]),
        ("medical", ["medical office building", "mob", "healthcare facility"]),
        ("industrial", ["industrial", "warehouse", "distribution center", "flex space",
                        "manufacturing", "logistics"]),
        ("retail", ["retail", "shopping center", "strip center", "strip mall", "shopping mall"]),
        ("mixed-use", ["mixed-use", "mixed use"]),
        ("multifamily", ["multifamily", "multi-family", "apartment complex", "garden style apartment"]),
        ("hospitality", ["hotel", "hospitality", "motel"]),
        ("self-storage", ["self-storage", "self storage", "storage facility"]),
        ("net lease", ["net lease", "single tenant net", "nnn lease"]),
        ("land", ["development site", "vacant land"]),
    ]
    for ptype, keywords in type_keywords_ordered:
        if any(kw in prop_text_lower for kw in keywords):
            details["property_type"] = ptype.title()
            break

    # ── Units / count ──
    unit_patterns = [
        r"(?:Number of Units|Unit Count|Total Units)[\s:]*(\d+)",
        r"(\d+)\s*(?:units?|suites?|spaces?)\s",
    ]
    for pattern in unit_patterns:
        units = re.search(pattern, property_text, re.IGNORECASE)
        if units:
            try:
                details["total_units"] = int(units.group(1))
            except ValueError:
                pass
            break

    # ── Total SF ──
    sf_patterns = [
        r"[±~+/-]*\s*(\d{1,3}(?:,\d{3})+)\s*(?:rentable\s*)?(?:square\s*foot|square\s*feet|SF|RSF|sq\.?\s*ft\.?)",
        r"(?:Total|Building|Rentable|Leasable|Net Rentable|GLA|NRA|Gross)?\s*(?:Square\s*F(?:ee|oo)t(?:age)?|SF|Sq\.?\s*Ft\.?|RSF|GLA|NRA)[\s:]*[±~]?\s*(\d{1,3}(?:,\d{3})*(?:\.\d+)?)",
    ]
    for pattern in sf_patterns:
        sf = extract_numeric_value(property_text, pattern)
        if sf and sf > 500:
            details["total_sf"] = sf
            break

    # ── Year built ──
    year_patterns = [
        r"(?:Year\s*Built|Built|Constructed|Year of Construction)[\s:]+(\d{4})",
        r"(?:built|constructed)\s+(?:in\s+)?(\d{4})",
    ]
    for pattern in year_patterns:
        year = re.search(pattern, text, re.IGNORECASE)
        if year:
            y = int(year.group(1))
            if 1900 <= y <= 2030:
                details["year_built"] = y
                break

    # ── Asking price ──
    amounts = find_all_dollar_amounts(text)
    price_keywords = ["asking", "price", "list", "offering", "purchase", "acquisition", "sale", "value"]
    for amt in amounts:
        if any(kw in amt["context"] for kw in price_keywords):
            if amt["value"] > 100000:
                details["asking_price"] = amt["value"]
                break

    # If no labeled price, look for loan-related amounts (common in financing memos)
    if "asking_price" not in details:
        loan_pattern = r"(?:Loan\s*Amount|Financing)[\s:]*\$?([\d,]+)"
        loan_match = re.search(loan_pattern, text[:10000], re.IGNORECASE)
        if loan_match:
            loan_amt = float(loan_match.group(1).replace(",", ""))
            if loan_amt > 1_000_000:
                # Estimate property value from loan at ~65% LTV
                details["asking_price"] = round(loan_amt / 0.65, 0)

    # Fallback: largest dollar amount > $1M
    if "asking_price" not in details and amounts:
        large_amounts = [a for a in amounts if a["value"] >= 1_000_000]
        if large_amounts:
            large_amounts.sort(key=lambda a: a["value"], reverse=True)
            skip_keywords = ["revenue", "expense", "income", "tax", "fee", "cost", "budget", "loan", "debt"]
            for amt in large_amounts:
                if not any(kw in amt["context"] for kw in skip_keywords):
                    details["asking_price"] = amt["value"]
                    break

    return details


def extract_financial_details(text: str) -> Dict[str, Any]:
    """Extract financial metrics with expanded pattern matching."""
    financials = {}

    # ── Cap rate ──
    cap_patterns = [
        r"[Cc]ap(?:italization)?\s*[Rr]ate[\s:]*(\d+\.?\d*)\s*%",
        r"(\d+\.?\d*)\s*%\s*[Cc]ap(?:italization)?\s*[Rr]ate",
        r"[Cc]ap[\s:]*(\d+\.?\d*)\s*%",
        r"going[\s-]in\s*(?:cap\s*)?(?:rate)?[\s:]*(\d+\.?\d*)\s*%",
    ]
    for pattern in cap_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            val = float(match.group(1))
            if 1 < val < 20:
                financials["cap_rate"] = val
                break

    # ── NOI ──
    noi_patterns = [
        r"(?:NOI|Net\s*Operating\s*Income)[\s:]*\$?([\d,]+(?:\.\d+)?)\s*(?:M|million)?",
        r"\$?([\d,]+(?:\.\d+)?)\s*(?:M|million)?\s*(?:NOI|Net\s*Operating\s*Income)",
    ]
    for pattern in noi_patterns:
        noi = extract_numeric_value(text, pattern)
        if noi:
            if noi < 1000:
                noi *= 1_000_000
            financials["noi"] = noi
            break

    # ── Revenue ──
    rev_patterns = [
        r"(?:Effective Gross Income|EGI)[\s:]*\$?([\d,]+(?:\.\d+)?)",
        r"(?:Annual|Gross|Total|Effective Gross)\s*(?:Revenue|Income|Rent|Rental Income)[\s:]*\$?([\d,]+(?:\.\d+)?)",
    ]
    for pattern in rev_patterns:
        rev = extract_numeric_value(text, pattern)
        if rev and rev > 10000:
            financials["annual_revenue"] = rev
            break

    # ── Operating expenses ── (MUST match "Total Operating Expenses" specifically)
    opex_matches = []
    for match in re.finditer(
        r"Total\s+Operating\s+Expenses?\s*\$?([\d,]+(?:\.\d+)?)",
        text, re.IGNORECASE
    ):
        try:
            val = float(match.group(1).replace(",", ""))
            if val > 50000:  # Must be at least $50K to be real OpEx
                opex_matches.append(val)
        except (ValueError, TypeError):
            pass

    if opex_matches:
        # Take the most common value or the first reasonable one
        financials["operating_expenses"] = opex_matches[0]

    # ── Occupancy / Vacancy ──
    occ_patterns = [
        r"(?:Physical\s*)?[Oo]ccupancy[\s:]*(\d+\.?\d*)\s*%",
        r"(\d+\.?\d*)\s*%\s*(?:occupied|occupancy|leased)",
    ]
    for pattern in occ_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            val = float(match.group(1))
            if 0 < val <= 100:
                financials["occupancy_rate"] = val
                financials["vacancy_rate"] = round(100 - val, 2)
                break

    if "vacancy_rate" not in financials:
        vac_patterns = [r"[Vv]acancy[\s:]*(\d+\.?\d*)\s*%"]
        for pattern in vac_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                val = float(match.group(1))
                if 0 <= val <= 100:
                    financials["vacancy_rate"] = val
                    financials["occupancy_rate"] = round(100 - val, 2)
                    break

    # ── DSCR ──
    dscr_pattern = r"(?:DSCR|Debt\s*Service\s*Coverage)[\s:]*(\d+\.?\d*)\s*x?"
    match = re.search(dscr_pattern, text, re.IGNORECASE)
    if match:
        val = float(match.group(1))
        if 0.5 < val < 5:
            financials["dscr"] = val

    # ── Derive NOI from revenue and expenses if missing ──
    if "noi" not in financials and "annual_revenue" in financials and "operating_expenses" in financials:
        financials["noi"] = financials["annual_revenue"] - financials["operating_expenses"]

    return financials


# ═══════════════════════════════════════════════════════════════
# PDF TABLE → RENT ROLL
# ═══════════════════════════════════════════════════════════════

def extract_rent_roll_from_pdf_tables(tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Try to extract rent roll data from PDF tables using fuzzy column matching."""
    rent_roll = []

    for table_info in tables:
        table_data = table_info.get("data", [])
        if not table_data or len(table_data) < 2:
            continue

        # Find header row using fuzzy matching
        header_row_idx = None
        best_col_map = {}

        for idx, row in enumerate(table_data[:8]):
            if not row:
                continue
            col_map = {}
            for ci, cell in enumerate(row):
                if cell is None:
                    continue
                field = fuzzy_match_column(str(cell))
                if field and field not in col_map.values():
                    col_map[ci] = field

            if len(col_map) >= 2 and len(col_map) > len(best_col_map):
                header_row_idx = idx
                best_col_map = col_map

        if header_row_idx is None or not best_col_map:
            continue

        # Parse data rows
        for row in table_data[header_row_idx + 1:]:
            if not row or not any(row):
                continue

            first_val = str(row[0] or '').lower().strip()
            if any(skip in first_val for skip in ['total', 'subtotal', 'sum', '---']):
                continue

            entry = {}
            for ci, val in enumerate(row):
                if ci not in best_col_map or val is None:
                    continue
                field = best_col_map[ci]
                val_str = str(val).strip()
                if not val_str or val_str.lower() in ["", "none", "-", "n/a"]:
                    continue

                if field in ["sf", "annual_rent", "monthly_rent", "rent_psf"]:
                    try:
                        cleaned = re.sub(r'[,$\s\(\)]', '', val_str)
                        if cleaned:
                            entry[field] = abs(float(cleaned))
                    except (ValueError, TypeError):
                        pass
                else:
                    entry[field] = val_str

            # Derive
            if "monthly_rent" in entry and "annual_rent" not in entry:
                entry["annual_rent"] = entry.pop("monthly_rent") * 12

            has_id = "unit" in entry or "tenant" in entry
            has_num = "sf" in entry or "annual_rent" in entry or "rent_psf" in entry
            if has_id and has_num:
                rent_roll.append(entry)

    return rent_roll


# ═══════════════════════════════════════════════════════════════
# MAIN PARSE FUNCTION
# ═══════════════════════════════════════════════════════════════

def parse_offering_memorandum(
    pdf_path: Optional[str] = None,
    excel_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Parse offering memorandum from PDF and/or Excel.
    Returns canonical OM structure with property, financials, rent_roll, assumptions, and raw_text.
    """
    result = {
        "property": {},
        "financials": {},
        "rent_roll": [],
        "assumptions": {
            "exit_cap_rate": None,
            "noi_growth": 3.0,
            "hold_period": 5,
            "ltv": 65.0,
            "interest_rate": 5.75,
            "amortization_years": 25,
        },
        "raw_text": "",
    }

    errors = []
    warnings = []

    # Extract from PDF
    if pdf_path:
        try:
            text, tables = extract_text_from_pdf(pdf_path)
            result["raw_text"] = text

            # Extract property details
            property_details = extract_property_details(text)
            result["property"].update(property_details)

            # Extract financials
            financials = extract_financial_details(text)
            result["financials"].update(financials)

            # Try to extract rent roll from PDF tables
            if tables:
                pdf_rent_roll = extract_rent_roll_from_pdf_tables(tables)
                if pdf_rent_roll:
                    result["rent_roll"] = pdf_rent_roll
                    warnings.append(f"Extracted {len(pdf_rent_roll)} rent roll entries from PDF tables")
                else:
                    warnings.append(f"Found {len(tables)} tables in PDF but no rent roll structure detected")
            else:
                warnings.append("No tables found in PDF")

            if not result["property"]:
                warnings.append("Could not extract property details from PDF")
            if not result["financials"]:
                warnings.append("Could not extract financial details from PDF")

        except Exception as e:
            errors.append(f"PDF parsing error: {str(e)}")
            import traceback
            traceback.print_exc()

    # Extract from Excel
    if excel_path:
        try:
            rent_roll = extract_excel_rent_roll(excel_path)
            if rent_roll:
                result["rent_roll"] = rent_roll
                warnings.append(f"Extracted {len(rent_roll)} rent roll entries from Excel")
            else:
                warnings.append("No rent roll data found in Excel — check column headers")
        except Exception as e:
            errors.append(f"Excel parsing error: {str(e)}")
            import traceback
            traceback.print_exc()

    # Smart defaults for assumptions
    cap_rate = result["financials"].get("cap_rate")
    if cap_rate:
        if cap_rate > 1:
            result["assumptions"]["exit_cap_rate"] = cap_rate + 0.25
        else:
            result["assumptions"]["exit_cap_rate"] = (cap_rate * 100) + 0.25

    # Add property total_sf to financials if missing
    if "total_sf" in result["property"] and "asking_price" in result["property"]:
        if result["property"]["total_sf"] > 0:
            result["financials"]["price_per_sf"] = round(
                result["property"]["asking_price"] / result["property"]["total_sf"], 2
            )

    # Move asking_price to financials too for consistency
    if "asking_price" in result["property"]:
        result["financials"].setdefault("asking_price", result["property"]["asking_price"])

    return {
        "parsed_data": result,
        "parsing_report": {
            "errors": errors,
            "warnings": warnings,
            "quality_score": calculate_quality_score(result),
        },
    }


def calculate_quality_score(parsed_data: Dict[str, Any]) -> float:
    """Calculate quality score of parsed data (0-100)."""
    score = 0.0

    weights = {
        "property_name": 10,
        "property_type": 5,
        "address": 10,
        "total_sf": 10,
        "asking_price": 15,
        "noi": 15,
        "cap_rate": 10,
        "rent_roll": 15,
        "occupancy": 5,
        "revenue": 5,
    }

    checks = {
        "property_name": bool(parsed_data.get("property", {}).get("name")),
        "property_type": bool(parsed_data.get("property", {}).get("property_type")),
        "address": bool(parsed_data.get("property", {}).get("address")),
        "total_sf": bool(parsed_data.get("property", {}).get("total_sf")),
        "asking_price": bool(parsed_data.get("property", {}).get("asking_price") or parsed_data.get("financials", {}).get("asking_price")),
        "noi": bool(parsed_data.get("financials", {}).get("noi")),
        "cap_rate": bool(parsed_data.get("financials", {}).get("cap_rate")),
        "rent_roll": bool(parsed_data.get("rent_roll")),
        "occupancy": bool(parsed_data.get("financials", {}).get("occupancy_rate")),
        "revenue": bool(parsed_data.get("financials", {}).get("annual_revenue")),
    }

    for field, weight in weights.items():
        if checks.get(field, False):
            score += weight

    return min(score, 100.0)
