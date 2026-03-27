"""
T12 (Trailing 12-Month Operating Statement) Parser

Parses Excel/CSV files containing monthly operating statements and extracts
revenue, expense, and NOI data into the CRE Lytic deal data model.

Supported formats:
  - Standard T12 with months as columns and line items as rows
  - Yardi / MRI / AppFolio property management exports
  - Annualized summary format (single-column totals)
"""

import re
import csv
import calendar
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# LINE ITEM ALIASES — fuzzy matching for common T12 row labels
# ═══════════════════════════════════════════════════════════════

LINE_ITEM_ALIASES: Dict[str, List[str]] = {
    "gross_potential_rent": [
        "gross potential rent", "gpr", "potential rental income",
        "scheduled rent", "gross scheduled rent", "gross rent potential",
        "potential rent", "scheduled base rent", "market rent",
        "gross potential revenue", "potential gross rent",
    ],
    "vacancy_loss": [
        "vacancy loss", "vacancy", "vacancy & credit loss",
        "vacancy and credit loss", "less vacancy", "vacancy allowance",
        "economic vacancy", "physical vacancy", "vac loss",
        "vacancy/credit loss", "less: vacancy", "credit loss",
    ],
    "concessions": [
        "concessions", "rent concessions", "lease concessions",
        "free rent", "rent abatement", "less: concessions",
    ],
    "bad_debt": [
        "bad debt", "bad debt expense", "uncollectible",
        "write-offs", "less: bad debt", "collection loss",
    ],
    "other_income": [
        "other income", "other revenue", "miscellaneous income",
        "misc income", "ancillary income", "sundry income",
        "non-rental income", "other operating income",
    ],
    "effective_gross_revenue": [
        "effective gross revenue", "egr", "effective gross income",
        "egi", "total revenue", "total income", "gross revenue",
        "net rental income", "total operating revenue",
        "effective revenue", "total effective revenue",
    ],
    "management_fee": [
        "management fee", "management", "property management",
        "management expense", "mgmt fee", "prop mgmt",
        "property management fee", "manager fee",
    ],
    "insurance": [
        "insurance", "property insurance", "hazard insurance",
        "liability insurance", "insurance expense", "ins",
        "casualty insurance", "general insurance",
    ],
    "taxes": [
        "taxes", "real estate taxes", "property taxes", "re taxes",
        "ad valorem taxes", "tax expense", "property tax",
        "real property taxes", "real estate tax",
    ],
    "repairs_maintenance": [
        "repairs maintenance", "repairs & maintenance", "r&m",
        "repairs and maintenance", "maintenance", "repairs",
        "building maintenance", "general maintenance",
        "maintenance & repairs", "building repairs",
    ],
    "utilities": [
        "utilities", "utility", "electric", "electricity",
        "gas", "water", "water & sewer", "water/sewer",
        "trash", "trash removal", "utility expense",
        "total utilities",
    ],
    "payroll": [
        "payroll", "payroll expense", "salaries", "wages",
        "salaries & wages", "employee expense", "personnel",
        "staff expense", "labor", "on-site payroll",
    ],
    "general_admin": [
        "general & administrative", "g&a", "general admin",
        "administrative", "admin expense", "office expense",
        "general and administrative", "g & a",
    ],
    "marketing": [
        "marketing", "advertising", "marketing & advertising",
        "leasing & marketing", "promotion", "marketing expense",
    ],
    "contract_services": [
        "contract services", "contracted services",
        "janitorial", "landscaping", "security", "cleaning",
        "pest control", "elevator maintenance",
    ],
    "other_opex": [
        "other opex", "other operating expense", "other expense",
        "miscellaneous expense", "misc expense", "sundry expense",
        "other operating expenses", "other expenses",
    ],
    "total_opex": [
        "total opex", "total operating expenses", "total expenses",
        "operating expenses", "total operating expense",
        "total expense", "total op expenses",
    ],
    "noi": [
        "noi", "net operating income", "net income",
        "operating income", "net oper income",
    ],
    "capex": [
        "capex", "capital expenditures", "capital expenses",
        "capital improvements", "cap ex", "capital reserves",
        "replacement reserves", "capital expenditure",
    ],
    "debt_service": [
        "debt service", "mortgage", "loan payment",
        "mortgage payment", "p&i", "principal & interest",
        "annual debt service", "total debt service",
    ],
}

# Month detection patterns
MONTH_PATTERNS: Dict[str, int] = {}
for i, name in enumerate(calendar.month_name[1:], 1):
    MONTH_PATTERNS[name.lower()] = i
    MONTH_PATTERNS[name[:3].lower()] = i
    MONTH_PATTERNS[f"m{i}"] = i
    MONTH_PATTERNS[f"month {i}"] = i
    MONTH_PATTERNS[f"month{i}"] = i

# Yardi/MRI/AppFolio sheet name indicators
YARDI_INDICATORS = [
    "financial overview", "income statement", "p&l", "profit and loss",
    "operating statement", "rent roll activity", "gl detail",
]
MRI_INDICATORS = [
    "mri", "financial statement", "income expense",
]
APPFOLIO_INDICATORS = [
    "appfolio", "owner statement", "profit and loss",
]


# ═══════════════════════════════════════════════════════════════
# UTILITY HELPERS
# ═══════════════════════════════════════════════════════════════

def _cell_val(ws, row: int, col: int) -> Any:
    """Get cell value, returning None for empty cells."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == "" or val.lower() == "nan":
            return None
    return val


def _to_float(val) -> float:
    """Convert a value to float, handling strings with commas, parens, currency symbols."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace("$", "").replace("%", "").strip()
    # Handle accounting-style negative: (1,234)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    # Handle trailing minus: 1,234-
    if s.endswith("-") and not s.startswith("-"):
        s = "-" + s[:-1]
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _fuzzy_match(text: str, aliases: List[str], threshold: float = 0.70) -> bool:
    """Check if text fuzzy-matches any alias above the given threshold."""
    text_lower = text.lower().strip()
    for alias in aliases:
        if alias in text_lower or text_lower in alias:
            return True
        ratio = SequenceMatcher(None, text_lower, alias).ratio()
        if ratio >= threshold:
            return True
    return False


def _identify_line_item(label: str) -> Optional[str]:
    """Identify which canonical line item a label corresponds to."""
    if not label:
        return None
    label_lower = label.lower().strip()
    # Direct / substring match first (faster)
    for canonical, aliases in LINE_ITEM_ALIASES.items():
        for alias in aliases:
            if alias == label_lower or alias in label_lower or label_lower in alias:
                return canonical
    # Fuzzy fallback
    for canonical, aliases in LINE_ITEM_ALIASES.items():
        if _fuzzy_match(label, aliases, threshold=0.75):
            return canonical
    return None


def _detect_month(header: str) -> Optional[int]:
    """Detect a month number (1-12) from a column header string."""
    if header is None:
        return None
    h = str(header).lower().strip()

    # Exact pattern match
    if h in MONTH_PATTERNS:
        return MONTH_PATTERNS[h]

    # Check if header starts with a known month name/abbreviation
    for pattern, month_num in MONTH_PATTERNS.items():
        if h.startswith(pattern) and len(pattern) >= 3:
            return month_num

    # Try parsing as a date string: "Jan 2024", "1/2024", "2024-01", etc.
    for fmt in ("%b %Y", "%B %Y", "%m/%Y", "%Y-%m", "%b-%Y", "%B-%y", "%b-%y"):
        try:
            dt = datetime.strptime(h, fmt)
            return dt.month
        except ValueError:
            continue

    return None


def _detect_month_year(header: str) -> Optional[Tuple[int, int]]:
    """Detect month and year from a column header. Returns (month, year) or None."""
    if header is None:
        return None
    h = str(header).strip()

    for fmt in ("%b %Y", "%B %Y", "%m/%Y", "%Y-%m", "%b-%Y", "%B-%y", "%b-%y",
                "%b %y", "%B %y", "%m/%y"):
        try:
            dt = datetime.strptime(h, fmt)
            year = dt.year
            # Handle 2-digit years
            if year < 100:
                year += 2000
            return (dt.month, year)
        except ValueError:
            continue

    # Fallback: month name only (no year)
    month = _detect_month(h)
    if month is not None:
        return (month, None)

    return None


# ═══════════════════════════════════════════════════════════════
# FORMAT DETECTION
# ═══════════════════════════════════════════════════════════════

def detect_t12_format(ws) -> str:
    """
    Detect the T12 format of a worksheet.

    Returns one of:
        - "monthly_columns" : standard T12 with months as column headers
        - "yardi"           : Yardi-style export
        - "mri"             : MRI-style export
        - "appfolio"        : AppFolio-style export
        - "summary"         : annualized summary (no monthly breakdown)
        - "unknown"         : could not determine format

    Scans the first 20 rows for month headers and vendor indicators.
    """
    month_col_count = 0
    has_annual_col = False
    vendor_hint = None

    max_scan_row = min(20, ws.max_row or 1)
    max_scan_col = min(30, ws.max_column or 1)

    for row in range(1, max_scan_row + 1):
        for col in range(1, max_scan_col + 1):
            val = _cell_val(ws, row, col)
            if val is None:
                continue
            val_str = str(val).lower().strip()

            # Check for vendor indicators
            for indicator in YARDI_INDICATORS:
                if indicator in val_str:
                    vendor_hint = "yardi"
            for indicator in MRI_INDICATORS:
                if indicator in val_str:
                    vendor_hint = "mri"
            for indicator in APPFOLIO_INDICATORS:
                if indicator in val_str:
                    vendor_hint = "appfolio"

            # Count month headers
            if _detect_month(val_str) is not None:
                month_col_count += 1

            # Check for annual/total column
            if val_str in ("annual", "total", "annualized", "ytd", "12 month total",
                           "trailing 12", "t12", "t-12"):
                has_annual_col = True

    if vendor_hint:
        logger.info("Detected vendor format: %s", vendor_hint)
        return vendor_hint

    if month_col_count >= 6:
        logger.info("Detected monthly_columns format (%d month headers found)", month_col_count)
        return "monthly_columns"

    if has_annual_col and month_col_count < 3:
        logger.info("Detected summary format (annual column, %d month headers)", month_col_count)
        return "summary"

    logger.warning("Could not determine T12 format; defaulting to monthly_columns")
    return "unknown"


# ═══════════════════════════════════════════════════════════════
# MONTHLY DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════

def _find_header_row(ws, max_rows: int = 25) -> Optional[int]:
    """Find the row containing month column headers."""
    max_scan_col = min(30, ws.max_column or 1)
    best_row = None
    best_count = 0

    for row in range(1, min(max_rows + 1, (ws.max_row or 1) + 1)):
        month_count = 0
        for col in range(1, max_scan_col + 1):
            val = _cell_val(ws, row, col)
            if val is not None and _detect_month(str(val)) is not None:
                month_count += 1
        if month_count > best_count:
            best_count = month_count
            best_row = row

    if best_count >= 3:
        return best_row
    return None


def _build_month_col_map(ws, header_row: int) -> Dict[str, int]:
    """
    Build a mapping of 'YYYY-MM' period keys to column indices from the header row.

    If year information is not present in headers, infers based on the most
    recent trailing 12 months.
    """
    max_scan_col = min(30, ws.max_column or 1)
    col_map: Dict[str, int] = {}
    month_cols: List[Tuple[int, int, Optional[int]]] = []  # (col, month, year)

    for col in range(1, max_scan_col + 1):
        val = _cell_val(ws, header_row, col)
        if val is None:
            continue
        result = _detect_month_year(str(val))
        if result is not None:
            month, year = result
            month_cols.append((col, month, year))

    if not month_cols:
        return {}

    # If years are present, use them directly
    has_years = any(y is not None for _, _, y in month_cols)
    if has_years:
        for col, month, year in month_cols:
            if year is None:
                # Infer year from neighbors
                known_years = [y for _, _, y in month_cols if y is not None]
                year = max(known_years) if known_years else datetime.now().year
            key = f"{year:04d}-{month:02d}"
            col_map[key] = col
    else:
        # No year info — assume trailing 12 ending at most recent month
        # Sort by column order and assign years
        month_cols.sort(key=lambda x: x[0])
        now = datetime.now()
        # Work backwards from the last column
        last_month = month_cols[-1][1]
        # Assume the last column is the most recent month
        end_year = now.year
        if last_month > now.month:
            end_year -= 1

        for col, month, _ in reversed(month_cols):
            year = end_year
            if month > last_month:
                year -= 1
            key = f"{year:04d}-{month:02d}"
            col_map[key] = col

    return col_map


def extract_monthly_data(ws, header_row: Optional[int] = None) -> Dict[str, Dict[str, float]]:
    """
    Extract month-by-month line item data from a worksheet.

    Returns a dict keyed by period string ('YYYY-MM'), where each value is
    a dict of canonical line item names to dollar amounts.

    Example:
        {
            "2024-01": {"gross_potential_rent": 100000, "vacancy_loss": -5000, ...},
            "2024-02": {...},
        }
    """
    if header_row is None:
        header_row = _find_header_row(ws)
    if header_row is None:
        logger.warning("No month header row found; cannot extract monthly data")
        return {}

    col_map = _build_month_col_map(ws, header_row)
    if not col_map:
        logger.warning("No month columns mapped from header row %d", header_row)
        return {}

    logger.info("Found %d month columns in header row %d", len(col_map), header_row)

    monthly: Dict[str, Dict[str, float]] = {period: {} for period in col_map}

    # Determine label column (usually column 1 or 2)
    label_col = 1
    # Check if column 1 has category groupings and column 2 has actual labels
    sample_labels_c1 = []
    sample_labels_c2 = []
    for row in range(header_row + 1, min(header_row + 20, (ws.max_row or header_row) + 1)):
        v1 = _cell_val(ws, row, 1)
        v2 = _cell_val(ws, row, 2)
        if v1 and isinstance(v1, str):
            sample_labels_c1.append(v1)
        if v2 and isinstance(v2, str):
            sample_labels_c2.append(v2)

    # If column 2 has more identifiable line items, use it
    c1_matches = sum(1 for l in sample_labels_c1 if _identify_line_item(l) is not None)
    c2_matches = sum(1 for l in sample_labels_c2 if _identify_line_item(l) is not None)
    if c2_matches > c1_matches:
        label_col = 2

    # Parse data rows
    for row in range(header_row + 1, (ws.max_row or header_row) + 1):
        label = _cell_val(ws, row, label_col)
        if not label or not isinstance(label, str):
            continue

        canonical = _identify_line_item(label)
        if canonical is None:
            continue

        for period, col in col_map.items():
            val = _cell_val(ws, row, col)
            if val is not None:
                amount = _to_float(val)
                # Only store non-zero values or known zero-possible items
                if amount != 0.0 or canonical in ("vacancy_loss", "concessions", "bad_debt"):
                    monthly[period][canonical] = amount

    # Log what we found
    all_items = set()
    for period_data in monthly.values():
        all_items.update(period_data.keys())
    logger.info("Extracted %d unique line items across %d months", len(all_items), len(monthly))

    return monthly


def _extract_summary_data(ws) -> Dict[str, float]:
    """
    Extract annualized line items from a summary-format T12.

    Scans for a label column and a corresponding value column (typically
    the first numeric column to the right of the label).
    """
    summary: Dict[str, float] = {}
    max_row = ws.max_row or 1

    for row in range(1, max_row + 1):
        label = _cell_val(ws, row, 1)
        if not label or not isinstance(label, str):
            continue

        canonical = _identify_line_item(label)
        if canonical is None:
            continue

        # Find the first numeric value to the right
        max_col = min(10, ws.max_column or 1)
        for col in range(2, max_col + 1):
            val = _cell_val(ws, row, col)
            if val is not None:
                amount = _to_float(val)
                if amount != 0.0 or canonical in ("vacancy_loss", "concessions", "bad_debt"):
                    summary[canonical] = amount
                    break

    logger.info("Extracted %d line items from summary format", len(summary))
    return summary


# ═══════════════════════════════════════════════════════════════
# ANNUALIZATION
# ═══════════════════════════════════════════════════════════════

def annualize_t12(monthly_data: Dict[str, Dict[str, float]]) -> Dict[str, float]:
    """
    Compute annual totals from monthly data by summing each line item
    across all available months.

    If fewer than 12 months are present, the totals reflect only the
    months available (no pro-rata scaling).

    Returns a dict of canonical line item names to annualized amounts.
    """
    annual: Dict[str, float] = {}

    for period_data in monthly_data.values():
        for item, amount in period_data.items():
            annual[item] = annual.get(item, 0.0) + amount

    # Round to avoid floating-point noise
    for key in annual:
        annual[key] = round(annual[key], 2)

    # Derive computed fields if not already present
    if "effective_gross_revenue" not in annual:
        gpr = annual.get("gross_potential_rent", 0.0)
        vac = annual.get("vacancy_loss", 0.0)
        conc = annual.get("concessions", 0.0)
        bad = annual.get("bad_debt", 0.0)
        other = annual.get("other_income", 0.0)
        egr = gpr + vac + conc + bad + other  # vacancy/concessions are typically negative
        if egr != 0.0:
            annual["effective_gross_revenue"] = round(egr, 2)

    if "total_opex" not in annual:
        opex_items = [
            "management_fee", "insurance", "taxes", "repairs_maintenance",
            "utilities", "payroll", "general_admin", "marketing",
            "contract_services", "other_opex",
        ]
        total = sum(annual.get(item, 0.0) for item in opex_items)
        if total != 0.0:
            annual["total_opex"] = round(total, 2)

    if "noi" not in annual:
        egr = annual.get("effective_gross_revenue", 0.0)
        opex = annual.get("total_opex", 0.0)
        if egr != 0.0:
            annual["noi"] = round(egr - opex, 2)

    return annual


# ═══════════════════════════════════════════════════════════════
# WORKSHEET SELECTION
# ═══════════════════════════════════════════════════════════════

def _find_t12_sheet(wb) -> Optional[Any]:
    """
    Find the best worksheet in a workbook for T12 data.

    Looks for sheets with names containing T12-related keywords,
    then falls back to the first sheet with recognizable month headers.
    """
    t12_keywords = [
        "t12", "t-12", "trailing", "income", "operating statement",
        "p&l", "profit", "loss", "financial", "income statement",
        "income & expense", "income expense",
    ]

    # Pass 1: keyword match on sheet names
    for name in wb.sheetnames:
        name_lower = name.lower().strip()
        for kw in t12_keywords:
            if kw in name_lower:
                logger.info("Selected sheet '%s' by keyword match '%s'", name, kw)
                return wb[name]

    # Pass 2: find a sheet with month headers
    for name in wb.sheetnames:
        ws = wb[name]
        if _find_header_row(ws, max_rows=15) is not None:
            logger.info("Selected sheet '%s' by month header detection", name)
            return ws

    # Pass 3: use first sheet
    if wb.sheetnames:
        logger.warning("No T12 sheet identified; using first sheet '%s'", wb.sheetnames[0])
        return wb[wb.sheetnames[0]]

    return None


# ═══════════════════════════════════════════════════════════════
# CSV SUPPORT
# ═══════════════════════════════════════════════════════════════

def _parse_t12_csv(file_path: str) -> Dict[str, Any]:
    """
    Parse a CSV-format T12 operating statement.

    Converts the CSV into an in-memory structure compatible with the
    openpyxl worksheet interface, then delegates to the standard
    extraction pipeline.
    """
    logger.info("Parsing T12 from CSV: %s", file_path)

    rows: List[List[Any]] = []
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for csv_row in reader:
            rows.append(csv_row)

    if not rows:
        logger.error("CSV file is empty: %s", file_path)
        return _empty_result("csv")

    # Build a lightweight wrapper mimicking openpyxl worksheet
    class _CsvSheet:
        """Minimal worksheet-like wrapper around CSV row data."""
        def __init__(self, data: List[List[Any]]):
            self._data = data
            self.max_row = len(data)
            self.max_column = max((len(r) for r in data), default=0)

        def cell(self, row: int, column: int):
            class _Cell:
                def __init__(self, value):
                    self.value = value
            r = row - 1  # 1-indexed
            c = column - 1
            if 0 <= r < len(self._data) and 0 <= c < len(self._data[r]):
                raw = self._data[r][c]
                if raw == "":
                    return _Cell(None)
                return _Cell(raw)
            return _Cell(None)

    ws = _CsvSheet(rows)
    return _extract_from_worksheet(ws)


# ═══════════════════════════════════════════════════════════════
# CORE EXTRACTION PIPELINE
# ═══════════════════════════════════════════════════════════════

def _empty_result(fmt: str = "unknown") -> Dict[str, Any]:
    """Return a valid but empty T12 result structure."""
    return {
        "monthly": {},
        "annual": {},
        "format": fmt,
        "months_parsed": 0,
        "period_start": None,
        "period_end": None,
    }


def _extract_from_worksheet(ws) -> Dict[str, Any]:
    """
    Core extraction logic shared by Excel and CSV paths.

    Detects the format, extracts monthly or summary data, computes
    annual totals, and returns the canonical T12 result dict.
    """
    fmt = detect_t12_format(ws)

    monthly_data: Dict[str, Dict[str, float]] = {}
    annual_data: Dict[str, float] = {}

    if fmt == "summary":
        annual_data = _extract_summary_data(ws)
        monthly_data = {}
    else:
        # Monthly extraction for all column-based formats
        monthly_data = extract_monthly_data(ws)
        if monthly_data:
            annual_data = annualize_t12(monthly_data)
        else:
            # Fallback: try summary extraction
            logger.info("Monthly extraction yielded no data; falling back to summary extraction")
            annual_data = _extract_summary_data(ws)
            fmt = "summary"

    # Determine period range
    sorted_periods = sorted(monthly_data.keys()) if monthly_data else []
    period_start = sorted_periods[0] if sorted_periods else None
    period_end = sorted_periods[-1] if sorted_periods else None
    months_parsed = len(sorted_periods)

    return {
        "monthly": monthly_data,
        "annual": annual_data,
        "format": fmt,
        "months_parsed": months_parsed,
        "period_start": period_start,
        "period_end": period_end,
    }


# ═══════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════

def parse_t12_excel(file_path: str) -> Dict[str, Any]:
    """
    Main entry point: parse a T12 operating statement from an Excel or CSV file.

    Supports .xlsx, .xls (via openpyxl), and .csv formats. Detects the T12
    layout automatically and extracts monthly revenue, expense, and NOI data.

    Args:
        file_path: Path to the Excel or CSV file.

    Returns:
        A dict with the following structure::

            {
                "monthly": {
                    "2024-01": {"gross_potential_rent": 100000, ...},
                    ...
                },
                "annual": {
                    "gross_potential_rent": 1200000,
                    "vacancy_loss": -60000,
                    "effective_gross_revenue": 1140000,
                    "management_fee": 45600,
                    ...
                    "noi": 914400,
                },
                "format": "monthly_columns",
                "months_parsed": 12,
                "period_start": "2024-01",
                "period_end": "2024-12",
            }
    """
    logger.info("Parsing T12 from: %s", file_path)

    path = Path(file_path)
    if not path.exists():
        logger.error("File not found: %s", file_path)
        return _empty_result()

    suffix = path.suffix.lower()

    # CSV path
    if suffix == ".csv":
        try:
            return _parse_t12_csv(file_path)
        except Exception as e:
            logger.exception("Failed to parse CSV T12: %s", e)
            return _empty_result("csv")

    # Excel path
    if suffix not in (".xlsx", ".xls", ".xlsm"):
        logger.error("Unsupported file type: %s", suffix)
        return _empty_result()

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.exception("Failed to open workbook: %s", e)
        return _empty_result()

    try:
        ws = _find_t12_sheet(wb)
        if ws is None:
            logger.error("No usable worksheet found in %s", file_path)
            return _empty_result()

        result = _extract_from_worksheet(ws)
    except Exception as e:
        logger.exception("Error extracting T12 data: %s", e)
        result = _empty_result()
    finally:
        wb.close()

    logger.info(
        "T12 parse complete — format=%s, months=%d, line_items=%d",
        result.get("format", "unknown"),
        result.get("months_parsed", 0),
        len(result.get("annual", {})),
    )

    return result


def is_t12_file(file_path: str) -> bool:
    """
    Quick heuristic check: does this file look like a T12 operating statement?

    Checks sheet names and scans for month headers and income/expense keywords.
    Returns True if the file is likely a T12.
    """
    try:
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".csv":
            with open(file_path, "r", encoding="utf-8-sig") as f:
                header = f.readline().lower()
                month_hits = sum(1 for m in calendar.month_abbr[1:] if m.lower() in header)
                return month_hits >= 3

        if suffix not in (".xlsx", ".xls", ".xlsm"):
            return False

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = [s.lower().strip() for s in wb.sheetnames]
        wb.close()

        t12_keywords = ["t12", "t-12", "trailing", "operating statement",
                        "income statement", "p&l"]
        for sn in sheet_names:
            for kw in t12_keywords:
                if kw in sn:
                    return True

        return False
    except Exception:
        return False
