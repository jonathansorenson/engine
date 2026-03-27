"""Automated tests for V2 Export endpoints — Tier 1 & 2 verification."""

import io
import sys
import os

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pytest
from fastapi.testclient import TestClient

# Minimal app import — just need the export router
from fastapi import FastAPI
from app.routes.export import router

app = FastAPI()
app.include_router(router)
client = TestClient(app)

# ═══════════════════════════════════════════════════════════
# SAMPLE TEST DATA
# ═══════════════════════════════════════════════════════════

SAMPLE_V2_PAYLOAD = {
    "v2_state": {
        "assumptions": {
            "name": "Test Deal - 123 Main St",
            "assetType": "Office",
            "address": "123 Main St, Denver CO",
            "sf": 50000,
            "purchasePrice": 10000000,
            "acqCostPct": 1.5,
            "closingCosts": 250000,
            "ltv": 65,
            "rate": 6.25,
            "amort": 30,
            "ioPeriod": 2,
            "isIO": True,
            "origFee": 1.0,
            "prepayPct": 1.0,
            "year1NOI": 700000,
            "rentGrowth": 3,
            "vacancy": 5,
            "exitCap": 6.5,
            "holdPeriod": 5,
            "showTILC": True,
            "useMarketLeasing": True,
            "marketRenewalProb": 75,
            "marketVacantMonths": 6,
            "marketFreeRentMonths": 2,
            "marketTINewPSF": 15,
            "marketTIRenewalPSF": 10,
            "marketLCNewPct": 6,
            "marketLCRenewalPct": 4,
        },
        "waterfall": {
            "lpPercent": 90,
            "gpPercent": 10,
            "prefReturn": 8,
            "tier1Split": 80,
            "tier1Thresh": 15,
            "tier2Split": 70,
            "tier2Thresh": 20,
            "tier3Split": 60,
            "catchUp": True,
        },
        "tenants": [
            {"name": "Tenant A - Acme Corp", "sf": 25000, "rentPSF": 18, "camPSF": 4,
             "start": "2024-01-01", "end": "2027-06-30", "escalPct": 3},
            {"name": "Tenant B - Beta LLC", "sf": 15000, "rentPSF": 22, "camPSF": 5,
             "start": "2024-06-01", "end": "2026-03-31", "escalPct": 2.5},
            {"name": "Tenant C - Gamma Inc", "sf": 10000, "rentPSF": 16, "camPSF": 3.5,
             "start": "2025-01-01", "end": "2029-12-31", "escalPct": 2},
        ],
        "valueAddEvents": [
            {"label": "Lobby Renovation Income", "year": 2, "amount": 50000, "type": "income"},
        ],
        "capexItems": [
            {"label": "HVAC Replacement", "year": 3, "amount": 200000},
        ],
    },
    "calc": {
        "pp": 10000000, "totalCost": 10400000, "loan": 6500000, "equity": 3900000,
        "origFee": 65000, "acqCost": 150000,
        "years": [
            {"yr": 1, "baseRent": 870000, "cam": 197500, "noi": 700000, "annDS": 490000,
             "vacancyLoss": 0, "freeRentLoss": 0, "capexRes": 50000, "specCapex": 0,
             "tiLC": 0, "vaInc": 0, "vaCost": 0, "levCF": 160000, "loanBal": 6400000},
            {"yr": 2, "baseRent": 896100, "cam": 203425, "noi": 721000, "annDS": 490000,
             "vacancyLoss": 50000, "freeRentLoss": 20000, "capexRes": 50000, "specCapex": 0,
             "tiLC": 30000, "vaInc": 50000, "vaCost": 0, "levCF": 131000, "loanBal": 6290000},
            {"yr": 3, "baseRent": 922983, "cam": 209528, "noi": 742630, "annDS": 490000,
             "vacancyLoss": 30000, "freeRentLoss": 10000, "capexRes": 50000, "specCapex": 200000,
             "tiLC": 0, "vaInc": 0, "vaCost": 0, "levCF": -37370, "loanBal": 6175000},
            {"yr": 4, "baseRent": 950672, "cam": 215814, "noi": 764909, "annDS": 490000,
             "vacancyLoss": 0, "freeRentLoss": 0, "capexRes": 50000, "specCapex": 0,
             "tiLC": 0, "vaInc": 0, "vaCost": 0, "levCF": 224909, "loanBal": 6055000},
            {"yr": 5, "baseRent": 979192, "cam": 222289, "noi": 787856, "annDS": 490000,
             "vacancyLoss": 0, "freeRentLoss": 0, "capexRes": 50000, "specCapex": 0,
             "tiLC": 0, "vaInc": 0, "vaCost": 0, "levCF": 247856, "loanBal": 5930000},
            {"yr": 6, "baseRent": 1008568, "cam": 228958, "noi": 811492, "annDS": 0,
             "vacancyLoss": 0, "freeRentLoss": 0, "capexRes": 0, "specCapex": 0,
             "tiLC": 0, "vaInc": 0, "vaCost": 0, "levCF": 0, "loanBal": 0},
        ],
        "holdYears": [],
        "exitNOI": 811492, "exitVal": 12484492, "exitBal": 5930000,
        "saleNet": 6495192, "prepayAmt": 59300,
        "goingCap": 7.0, "yoc": 6.73, "levIRR": 22.5, "unlevIRR": 12.1,
        "em": 2.35, "dscr": 1.43, "avgCoC": 4.48,
        "levCFsForIRR": [-3900000, 160000, 131000, -37370, 224909, 6743048],
        "totalReturn": 7221587,
        "lpEq": 3510000, "gpEq": 390000,
        "lpOut": 6200000, "gpOut": 1021587,
        "lpIRR": 19.8, "gpIRR": 45.2, "lpEM": 1.77, "gpPromote": 631587,
        "goGreen": True,
    },
}

SAMPLE_V1_PAYLOAD = {
    "property_name": "V1 Test Property",
    "address": "456 Oak Ave",
    "property_type": "Retail",
    "asking_price": 5000000,
    "noi": 350000,
    "cap_rate": 7.0,
    "rentable_sf": 30000,
    "assumptions": {
        "ltv": 65,
        "interest_rate": 6.0,
        "exit_cap_rate": 7.5,
        "noi_growth": 2.0,
        "hold_period": 5,
        "amortization_years": 25,
    },
}


# ═══════════════════════════════════════════════════════════
# PHASE 1: API ENDPOINT TESTS
# ═══════════════════════════════════════════════════════════

class TestV1Export:
    """V1 export endpoint — regression tests."""

    def test_v1_export_returns_xlsx(self):
        r = client.post("/api/v1/export", json=SAMPLE_V1_PAYLOAD)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert len(r.content) > 1000

    def test_v1_export_empty_payload(self):
        r = client.post("/api/v1/export", json={})
        assert r.status_code == 200  # Should handle defaults gracefully


class TestV2Export:
    """V2 export endpoint — Tier 2 feature tests."""

    def test_v2_export_returns_xlsx(self):
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert len(r.content) > 5000  # Should be substantial

    def test_v2_export_has_all_sheets(self):
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        expected_sheets = ["Assumptions", "Cash Flows", "Returns", "Waterfall",
                          "Value-Add & CapEx", "Rent Roll"]
        for sheet in expected_sheets:
            assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"

    def test_v2_export_has_charts_sheet(self):
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        assert "Charts" in wb.sheetnames, "Charts sheet missing"
        charts_ws = wb["Charts"]
        assert len(charts_ws._charts) >= 2, f"Expected >=2 charts, got {len(charts_ws._charts)}"

    def test_v2_export_has_key_insights_sheet(self):
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        assert "Key Insights" in wb.sheetnames, "Key Insights sheet missing"

    def test_v2_export_has_formulas_in_cash_flows(self):
        """KEY TEST: Excel must contain formulas, not just static values."""
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        cf_ws = wb["Cash Flows"]

        # Find NCF row
        ncf_row = None
        for row in cf_ws.iter_rows(min_col=1, max_col=1):
            val = row[0].value
            if val and "Net Cash Flow" in str(val):
                ncf_row = row[0].row
                break
        assert ncf_row is not None, "Net Cash Flow row not found in Cash Flows sheet"

        # Cell in column B should be a formula
        ncf_cell = cf_ws.cell(row=ncf_row, column=2).value
        assert isinstance(ncf_cell, str) and ncf_cell.startswith("="), \
            f"NCF cell B{ncf_row} should be a formula, got: {ncf_cell}"

    def test_v2_export_has_total_column_formulas(self):
        """Total column should use SUM formulas."""
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        cf_ws = wb["Cash Flows"]

        # Find the Total column (last data column)
        # The total column should have SUM formulas
        found_sum = False
        for row in cf_ws.iter_rows(min_row=4, max_row=cf_ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "SUM(" in cell.value:
                    found_sum = True
                    break
            if found_sum:
                break
        assert found_sum, "No SUM() formulas found in Cash Flows sheet"

    def test_v2_export_has_named_ranges(self):
        """Named ranges should exist for key assumptions."""
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        defined_names = list(wb.defined_names)
        for name in ["PurchasePrice", "LTV", "Rate"]:
            assert name in defined_names, f"Missing named range: {name}"

    def test_v2_export_has_deal_summary(self):
        """Returns sheet should have auto-generated deal summary."""
        from openpyxl import load_workbook
        r = client.post("/api/v1/export/v2", json=SAMPLE_V2_PAYLOAD)
        wb = load_workbook(io.BytesIO(r.content))
        returns_ws = wb["Returns"]
        # Scan for summary text
        found_summary = False
        for row in returns_ws.iter_rows(min_col=1, max_col=1):
            val = row[0].value
            if val and ("123 Main St" in str(val) or "Office" in str(val) or "50,000" in str(val)):
                found_summary = True
                break
        assert found_summary, "Deal summary paragraph not found in Returns sheet"

    def test_v2_export_empty_tenants(self):
        """Export should work with no tenants."""
        payload = {
            "v2_state": {
                "assumptions": SAMPLE_V2_PAYLOAD["v2_state"]["assumptions"],
                "waterfall": SAMPLE_V2_PAYLOAD["v2_state"]["waterfall"],
                "tenants": [],
                "valueAddEvents": [],
                "capexItems": [],
            },
            "calc": SAMPLE_V2_PAYLOAD["calc"],
        }
        r = client.post("/api/v1/export/v2", json=payload)
        assert r.status_code == 200

    def test_v2_export_empty_calc(self):
        """Export should handle empty calc gracefully."""
        payload = {
            "v2_state": SAMPLE_V2_PAYLOAD["v2_state"],
            "calc": {},
        }
        r = client.post("/api/v1/export/v2", json=payload)
        assert r.status_code == 200


class TestMemoExport:
    """Memo export endpoints — Tier 2 feature tests."""

    def test_memo_html_returns_html(self):
        r = client.post("/api/v1/export/v2/memo/html", json=SAMPLE_V2_PAYLOAD)
        assert r.status_code == 200
        assert "text/html" in r.headers["content-type"]
        content = r.content.decode('utf-8')
        assert "CRE" in content or "Deal Memo" in content

    def test_memo_html_has_executive_summary(self):
        r = client.post("/api/v1/export/v2/memo/html", json=SAMPLE_V2_PAYLOAD)
        content = r.content.decode('utf-8')
        assert "Executive Summary" in content or "Summary" in content

    def test_memo_html_has_cash_flow_table(self):
        r = client.post("/api/v1/export/v2/memo/html", json=SAMPLE_V2_PAYLOAD)
        content = r.content.decode('utf-8')
        assert "<table" in content
        assert "NOI" in content

    def test_memo_docx_returns_docx(self):
        r = client.post("/api/v1/export/v2/memo/docx", json=SAMPLE_V2_PAYLOAD)
        if r.status_code == 501:
            pytest.skip("python-docx not installed")
        assert r.status_code == 200
        assert "officedocument" in r.headers["content-type"]
        assert len(r.content) > 1000

    def test_memo_docx_empty_payload(self):
        r = client.post("/api/v1/export/v2/memo/docx", json={"v2_state": {}, "calc": {}})
        # Should not crash even with empty data
        assert r.status_code in (200, 501)


class TestEdgeCases:
    """Edge case and regression tests."""

    def test_zero_purchase_price(self):
        """No division by zero when PP=0."""
        payload = {
            "v2_state": {
                "assumptions": {**SAMPLE_V2_PAYLOAD["v2_state"]["assumptions"], "purchasePrice": 0},
                "waterfall": SAMPLE_V2_PAYLOAD["v2_state"]["waterfall"],
                "tenants": [], "valueAddEvents": [], "capexItems": [],
            },
            "calc": {**SAMPLE_V2_PAYLOAD["calc"], "pp": 0, "equity": 0, "loan": 0},
        }
        r = client.post("/api/v1/export/v2", json=payload)
        assert r.status_code == 200

    def test_single_year_hold(self):
        """Hold period = 1 should still work."""
        payload = {
            "v2_state": {
                "assumptions": {**SAMPLE_V2_PAYLOAD["v2_state"]["assumptions"], "holdPeriod": 1},
                "waterfall": SAMPLE_V2_PAYLOAD["v2_state"]["waterfall"],
                "tenants": [], "valueAddEvents": [], "capexItems": [],
            },
            "calc": {
                **SAMPLE_V2_PAYLOAD["calc"],
                "years": [SAMPLE_V2_PAYLOAD["calc"]["years"][0]],
            },
        }
        r = client.post("/api/v1/export/v2", json=payload)
        assert r.status_code == 200

    def test_no_waterfall_data(self):
        """Missing waterfall config should not crash."""
        payload = {
            "v2_state": {
                "assumptions": SAMPLE_V2_PAYLOAD["v2_state"]["assumptions"],
                "waterfall": {},
                "tenants": [], "valueAddEvents": [], "capexItems": [],
            },
            "calc": SAMPLE_V2_PAYLOAD["calc"],
        }
        r = client.post("/api/v1/export/v2", json=payload)
        assert r.status_code == 200
